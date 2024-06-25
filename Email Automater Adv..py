# dftlozwxupycipab
from email.message import EmailMessage
import ssl
import smtplib
from tkinter.filedialog import *
from tkinter import *
from openpyxl import load_workbook
import requests
import time
import random
win = Tk()
win.geometry("900x600")

# https://isitarealemail.com/api/email/validate?email=foo@bar.com


def SendMail():
    try:
        global EmailBody, EmailPerId, ReciverLastNo, ReciverFirstNo, PassLastNo, PassFirstNo, IdLastNo, IdFirstNo
        ifn = int(IdFirstNo.get())
        iln = int(IdLastNo.get())
        pfn = int(PassFirstNo.get())
        pln = int(PassLastNo.get())
        rfn = int(ReciverFirstNo.get())
        rln = int(ReciverLastNo.get())
        rfn -= 1
        ifn -= 1
        pfn -= 1
        EmailReciver = ReciverList[rfn:rln]
        EmailSender = IdList[ifn:iln]
        EmailPassword = PassList[pfn:pln]
        Esb = EmailSubject.get()
        EmailBody = EmailBody.get("1.0", 'end-1c')
        EmPerID = int(EmailPerId.get())
        print(EmPerID)
        Epi = EmPerID
        emsend = 0
        for i, j in zip(EmailSender, EmailPassword):
            for k in EmailReciver[Epi-EmPerID:Epi]:
                digits = "1234567890"
                Sn = i
                SenderN = Sn.replace("@gmail.com", "")
                SenderName = ""
                sub = k.partition("@")[2]
                if sub == "gmail.com":
                    sub = Esb
                for m in SenderN:
                    if m not in digits:
                        SenderName += m
                Message = EmailBody + "\n" + SenderName.strip()
                Email = EmailMessage()
                Email["From"] = i
                Email["To"] = k
                Email["Subject"] = sub
                Email.set_content(Message)
                context = ssl.create_default_context()
                with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
                    smtp.login(i, j)
                    smtp.sendmail(i, k, Email.as_string())
                    emsend += 1
                    print("Mail send to ", k, "     Mail No. ", emsend)
                time.sleep(random.randint(1, 20))
            Epi += EmPerID
    except ValueError:
        print("Please Fill All the Required Feilds")


def SelectReciverFile():
    global ReciverList
    ReciverFile = askopenfilename()
    if len(ReciverFile) == 0:
        pass
    else:
        Label(win, text=ReciverFile).place(x=15, y=193)
        LoadReciver = load_workbook(ReciverFile)
        ReciFile = LoadReciver.active
        ReciverList = list()
        for i in ReciFile['A']:
            ReciverList.append(i.value)


def SelectIdFile():
    global PassList, IdList
    ExcelFile = askopenfilename()
    if len(ExcelFile) == 0:
        pass
    else:
        Label(win, text=ExcelFile).place(x=15, y=40)
        LoadExcel = load_workbook(ExcelFile)
        IdFile = LoadExcel.active
        IdList = list()
        for i in IdFile['A']:
            IdList.append(i.value)
        PassList = list()
        for i in IdFile['B']:
            PassList.append(i.value)


# select id file
IdSelectBtn = Button(win, text="Select Email File", width=15,
                     command=SelectIdFile).place(x=15, y=15)
# id line first
IdFirstNo = Entry(win, width=10)
IdFirstNo.place(x=105, y=72)
IdFirstNoLabel = Label(win, text="ID First Row:").place(x=15, y=70)
# id line last
IdLastNo = Entry(win, width=10)
IdLastNo.place(x=105, y=92)
IdLastNoLabel = Label(win, text="ID Last Row:").place(x=15, y=90)
# pass first line
PassFirstNoLabel = Label(win, text="Pass First Row:").place(x=15, y=112)
PassFirstNo = Entry(win, width=10)
PassFirstNo.place(x=105, y=115)
# pass last line
PassLastNoLabel = Label(win, text="Pass Last Row:").place(x=15, y=132)
PassLastNo = Entry(win, width=10)
PassLastNo.place(x=105, y=135)
# select reciver file
ReciverSelectBtn = Button(win, text="Select Reciver File", width=15,
                          command=SelectReciverFile).place(x=15, y=165)
# reciver first line
ReciverFirstNoLabel = Label(win, text="Reciver First Row:").place(x=15, y=222)
ReciverFirstNo = Entry(win, width=10)
ReciverFirstNo.place(x=115, y=222)
# reciver first line
ReciverLastNoLabel = Label(win, text="Reciver Last Row:").place(x=15, y=242)
ReciverLastNo = Entry(win, width=10)
ReciverLastNo.place(x=115, y=245)
# email per id
EmailPerIdLabel = Label(win, text="Email Per ID:").place(x=15, y=262)
EmailPerId = Entry(win, width=10)
EmailPerId.place(x=115, y=265)
# email subject
EmailSubjectLabel = Label(win, text="Email Subject:").place(x=15, y=287)
EmailSubject = Entry(win, width=60)
EmailSubject.place(x=115, y=290)
# email body
EmailBodyLabel = Label(win, text="Email Body:").place(x=15, y=307)
EmailBody = Text(win, height=10, width=80)
EmailBody.place(x=115, y=310)
SendMailBtn = Button(win, text="SEND EMAIL", width=15,
                     command=SendMail).place(x=15, y=490)

win.mainloop()
